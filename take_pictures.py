import os
import pdb
import time
import openpyxl
import pyautogui
import subprocess

import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def switch_to_next_sheet():
    pyautogui.hotkey('ctrl', 'pagedown') 
    time.sleep(2)


def take_screenshot(output_image_path):
    time.sleep(2)
    screenshot = pyautogui.screenshot()
    screenshot.save(output_image_path)
    time.sleep(2)


def close_excel():
    os.system("taskkill /f /im excel.exe")
    print("Excel terminated using taskkill")
    time.sleep(2)


def check_existing_screenshot(file_name, output_directory, num_sheets):
    # Check if the screenshot already exists
    file_name = os.path.splitext(file_name)[0]
    cnt = 0
    for name in os.listdir(output_directory):
        if name.split('-')[0] == (file_name):
            cnt += 1
        if cnt == num_sheets:
            return True
    return False


def process_multiple_sheets(file_path, output_directory):
    # Load the Excel file to get the sheet names
    workbook = openpyxl.load_workbook(file_path)
    sheet_names = workbook.sheetnames
   
    process = subprocess.Popen(["start", "excel", file_path], shell=True)
    time.sleep(8) # Wait for Excel to open

    for sheet_name in sheet_names:
        output_image_path = os.path.join(output_directory, f"{os.path.splitext(os.path.basename(file_path))[0]}-{sheet_name}.png")
        # Open Excel and switch to the current sheet
        if os.path.exists(output_image_path):
            continue
        # Take a screenshot of the current sheet
        take_screenshot(output_image_path)
        switch_to_next_sheet()
        
    # Close the Excel file
    close_excel()
    process.kill()


def process_multiple_excel_files(directory_path, output_directory):
    # Create output directory if it doesn't exist
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    cnt = 0
    threshold = 9999
    # Loop through each Excel file in the directory
    for filename in os.listdir(directory_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(directory_path, filename)
            workbook = openpyxl.load_workbook(file_path)
            num_sheets = len(workbook.sheetnames)
            if check_existing_screenshot(filename, output_directory, num_sheets):
                print(f"Skipping {filename}")
                continue
            cnt += 1
            if cnt >= threshold:
                break
            process_multiple_sheets(file_path, output_directory)


if __name__ == '__main__':
    target_name = 'employee-recognition'
    directory_path = f'storage\{target_name}'
    output_directory = f'figures\{target_name}'
    process_multiple_excel_files(directory_path, output_directory)
