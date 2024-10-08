import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from tabulate import tabulate

import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def CSVtoSummaryTableExcel():
    df = pd.read_csv('test.csv', index_col=0, header=None)
    df.columns = ['Name', 'Url', 'Location', 'Tag']
    df.to_excel('tableSummary.xlsx', index=False)


def BeautifySummaryTable():
    # Load the Excel workbook and select the active worksheet
    workbook = load_workbook('tableSummary.xlsx')
    sheet = workbook.active

    # Adjust the column width to fit the contents
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Change the font for a specific range of cells (e.g., A1:C10)
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.font = Font(name='等线', size=12, bold=False, italic=False)

    # Save the modified workbook
    workbook.save('tableSummary.xlsx')



if __name__ == '__main__':

    data = pd.read_excel('tableSummary.xlsx')[['Name', "Tag", "Location"]]
    
    data['NameNew'] = data['Location'].apply(lambda x: x.split('/')[-1])
    data['NameNew'] = data['NameNew'].apply(lambda x: x.split('\\')[-1])

    data = data[['NameNew', "Tag", "Location"]]
    data.columns = ['Name', 'Tag', 'Location']

    print(tabulate(data, headers='keys', tablefmt='psql')) 

    data.to_excel('tableSummary.xlsx', index=False)

    BeautifySummaryTable()
